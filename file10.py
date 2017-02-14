from openpyxl import load_workbook

# Opening workbook to read

# -------------------------------------------------------------------------------------------------------------------
#                                             GENERAL ERRORS
# -------------------------------------------------------------------------------------------------------------------

fil=load_workbook(filename='error-code-ref-complete.xlsx')
# This script reads the xlsx file named error-code-ref-complete that is required to be stored
# in the same folder as this script file

sheet_ranges = fil['error-code-ref-general']
# Selecting the sheet in workbook to read

htmlfile=open('error-code-ref-general.html','w+',encoding="utf8")
# The HTML file generated will be named error-code-ref-general.html

htmlfile.write('<html>') # write HTML tag in the file
htmlfile.write('\n')
htmlfile.write('<head>')
htmlfile.write('\n')
htmlfile.write('</head>')
htmlfile.write('\n')
htmlfile.write('<style>') # Writing CSS for the table
htmlfile.write('\n')
htmlfile.write('table,tr,td')
htmlfile.write('\n')
htmlfile.write('{ border: 1px solid black; }') # Create a border around the table
htmlfile.write('\n')
htmlfile.write('</style>')
htmlfile.write('\n')
htmlfile.write('<body>') # We start the body tag in HTML file
htmlfile.write('\n')
i=1
while i<=10150:
    # Check for headings in the excel file
    if sheet_ranges['A'+str(i)].value !="Cause" \
            and sheet_ranges['A'+str(i)].value != "Type" \
            and sheet_ranges['A'+str(i)].value != "Action" \
            and sheet_ranges['A'+str(i)].value != "Code":

        htmlfile.write('<h2>') # We print the heading with tag h2
        htmlfile.write(sheet_ranges['A'+str(i)].value)
        htmlfile.write('</h2>')
        htmlfile.write('\n')
        i = i + 1
    else:
        htmlfile.write('<table>') #Creating table
        htmlfile.write('\n')
        htmlfile.write('<tr>') # Each row contains 2 cells
        htmlfile.write('\n')
        htmlfile.write('<td>')
        htmlfile.write('\n')
        htmlfile.write('Code')
        htmlfile.write('\n')
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        htmlfile.write('\n')
        val = sheet_ranges['B' + str(i)].value
        if val is None:
              htmlfile.write(' ')
        else:
              htmlfile.write(val)
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('</tr>')
        htmlfile.write('\n')

        htmlfile.write('<tr>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        htmlfile.write('\n')
        htmlfile.write('Type')
        htmlfile.write('\n')
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        val = sheet_ranges['B' + str(i + 1)].value
        if val is None:
              htmlfile.write(' ')
        else:
              htmlfile.write(val)
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('</tr>')
        htmlfile.write('\n')

        htmlfile.write('<tr>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        htmlfile.write('\n')
        htmlfile.write('Cause')
        htmlfile.write('\n')
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        val = sheet_ranges['B' + str(i+2)].value
        if val is None:
            htmlfile.write(' ')
        else:
            htmlfile.write(val)
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('</tr>')
        htmlfile.write('\n')

        htmlfile.write('<tr>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        htmlfile.write('\n')
        htmlfile.write('Action')
        htmlfile.write('\n')
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        val = sheet_ranges['B' + str(i+3)].value
        if val is None:
            htmlfile.write(' ')
        else:
            htmlfile.write(val)
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('</tr>')
        htmlfile.write('\n')
        htmlfile.write('\n')
        htmlfile.write('</table>')
        htmlfile.write('</br>')
        i = i + 4
htmlfile.write('\n')
htmlfile.write('</body>')
htmlfile.write('\n')
htmlfile.write('</html>')
htmlfile.close()

# -------------------------------------------------------------------------------------------------------------------
#                                             UPGRADE ERRORS
# -------------------------------------------------------------------------------------------------------------------

fil=load_workbook(filename='error-code-ref-complete.xlsx')

sheet_ranges = fil['error-code-ref-upgrade']
# Selecting the sheet in workbook to read
htmlfile=open('error-code-ref-upgrade.html','w+',encoding="utf8")
# HTML file generated error-code-ref-upgrade.html
htmlfile.write('<html>') # write HTML tag in the file
htmlfile.write('\n')
htmlfile.write('<head>')
htmlfile.write('\n')
htmlfile.write('</head>')
htmlfile.write('\n')
htmlfile.write('<style>') # Writing CSS for the table
htmlfile.write('\n')
htmlfile.write('table,tr,td')
htmlfile.write('\n')
htmlfile.write('{ border: 1px solid black; }') # Create a border around the table
htmlfile.write('\n')
htmlfile.write('</style>')
htmlfile.write('\n')
htmlfile.write('<body>') # Start body tag of HTML file
htmlfile.write('\n')
i=1
while i <= 1753:
    if sheet_ranges['A' + str(i)].value != "Cause" \
            and sheet_ranges['A' + str(i)].value != "Type" \
            and sheet_ranges['A' + str(i)].value != "Action" \
            and sheet_ranges['A' + str(i)].value != "Code":
        # Checking for headings in the Excel file
        htmlfile.write('<h2>') # We print the heading with tag h2
        htmlfile.write(sheet_ranges['A' + str(i)].value)
        htmlfile.write('</h2>')
        htmlfile.write('\n')
        i = i + 1
    else:
        htmlfile.write('<table>')  # Creating table
        htmlfile.write('\n')
        htmlfile.write('<tr>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        htmlfile.write('\n')
        htmlfile.write('Code')
        htmlfile.write('\n')
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        htmlfile.write('\n')
        val = sheet_ranges['B' + str(i)].value
        if val is None:
            htmlfile.write(' ')
        else:
            htmlfile.write(val)
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('</tr>')
        htmlfile.write('\n')

        htmlfile.write('<tr>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        htmlfile.write('\n')
        htmlfile.write('Type')
        htmlfile.write('\n')
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        val = sheet_ranges['B' + str(i + 1)].value
        if val is None:
            htmlfile.write(' ')
        else:
            htmlfile.write(val)
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('</tr>')
        htmlfile.write('\n')

        htmlfile.write('<tr>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        htmlfile.write('\n')
        htmlfile.write('Cause')
        htmlfile.write('\n')
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        val = sheet_ranges['B' + str(i + 2)].value
        if val is None:
            htmlfile.write(' ')
        else:
            htmlfile.write(val)
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('</tr>')
        htmlfile.write('\n')

        htmlfile.write('<tr>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        htmlfile.write('\n')
        htmlfile.write('Action')
        htmlfile.write('\n')
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('<td>')
        val = sheet_ranges['B' + str(i + 3)].value
        if val is None:
            htmlfile.write(' ')
        else:
            htmlfile.write(val)
        htmlfile.write('</td>')
        htmlfile.write('\n')
        htmlfile.write('</tr>')
        htmlfile.write('\n')
        htmlfile.write('\n')
        htmlfile.write('</table>')
        htmlfile.write('</br>')
        i = i + 4

htmlfile.write('\n')
htmlfile.write('</body>') # Close the body tag
htmlfile.write('\n')
htmlfile.write('</html>') # Close the html tag
htmlfile.close() # We close the HTML file

# --------------------------------Program ends here--------------------------------